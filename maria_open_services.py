import urllib.request, urllib.parse, json, base64, time, re
from collections import defaultdict

from insyte_env import EMAIL, KEY
BASE  = "https://api.myinsyte.com.au/v2"
AUTH  = "Basic " + base64.b64encode(f"{EMAIL}:{KEY}".encode()).decode()

def _get(path, retries=5):
    url = f"{BASE}/{path.lstrip('/')}"
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers={"Authorization": AUTH, "Accept": "application/json"})
            with urllib.request.urlopen(req, timeout=120) as r:
                return json.loads(r.read())
        except Exception as e:
            if attempt < retries - 1: time.sleep(2 ** attempt)
            else: raise

def fetch_all(path, label):
    PAGE, rows, skip, pg = 500, [], 0, 1
    sep = "&" if "?" in path else "?"
    while True:
        print(f"  {label} p{pg} ({len(rows)})...", flush=True)
        data = _get(f"{path}{sep}$top={PAGE}&$skip={skip}")
        page = data.get("value", data if isinstance(data, list) else [])
        rows.extend(page)
        if len(page) < PAGE: break
        skip += PAGE; pg += 1
    return rows

# 1. Find Maria Rundle
print("Finding Maria Rundle...")
users = fetch_all("/Users", "users")
maria = next((u for u in users if "rundle" in (u.get("FullName") or "").lower() or
              ("rundle" in (u.get("LastName") or "").lower())), None)
if not maria:
    # Try partial match
    maria = next((u for u in users if "maria" in (u.get("FirstName") or "").lower() and
                  "run" in (u.get("LastName") or "").lower()), None)

if not maria:
    print("Maria Rundle not found. All users:")
    for u in users:
        fn = u.get("FullName") or f"{u.get('FirstName','')} {u.get('LastName','')}".strip()
        print(f"  ID {u['ID']}: {fn}")
    exit(1)

maria_id = maria["ID"]
maria_name = maria.get("FullName") or f"{maria.get('FirstName','')} {maria.get('LastName','')}".strip()
print(f"  Found: {maria_name} (ID: {maria_id})\n")

# 2. Fetch all service jobs for Maria that are open (not closed/cancelled)
print("Fetching Maria's open service jobs...")
sf = urllib.parse.quote(
    f"SalesRepID eq {maria_id} and JobType eq 'type_job_service' and Status eq 'status_job_open'"
)
service_jobs = fetch_all(f"/Jobs?$filter={sf}", "service jobs")
print(f"  {len(service_jobs)} open service jobs found\n")

if not service_jobs:
    # Try without status filter to see what's there
    sf2 = urllib.parse.quote(f"SalesRepID eq {maria_id} and JobType eq 'type_job_service'")
    all_svc = fetch_all(f"/Jobs?$filter={sf2}", "all service jobs")
    print(f"  All service jobs (any status): {len(all_svc)}")
    from collections import Counter
    print(Counter(j.get("Status") for j in all_svc))
    service_jobs = [j for j in all_svc if j.get("Status") not in {"status_job_closed", "status_job_cancelled"}]
    print(f"  After filtering closed/cancelled: {len(service_jobs)}")

job_ids = [j["ID"] for j in service_jobs]
print(f"Job IDs to look up: {len(job_ids)}")

# 3. Fetch job lines for these service jobs to get line details + ETA
print("Fetching job lines for service jobs...")
CHUNK = 15
job_lines = defaultdict(list)
for i in range(0, len(job_ids), CHUNK):
    chunk = job_ids[i:i+CHUNK]
    fstr = urllib.parse.quote(" or ".join(f"JobID eq {c}" for c in chunk))
    try:
        data = _get(f"/JobLines?$filter={fstr}&$top={CHUNK*10}")
        for l in data.get("value", []):
            job_lines[l["JobID"]].append(l)
    except Exception as e:
        print(f"  Warning chunk {i}: {e}")
    if i % 300 == 0:
        print(f"  lines {min(i+CHUNK,len(job_ids))}/{len(job_ids)}...", flush=True)

# 4. Fetch CM Booking activities for these jobs (to find booking dates)
print("Fetching booking activities for these jobs...")
booking_dates = {}  # job_id -> earliest future booking date
WAVE = 15
for i in range(0, len(job_ids), WAVE):
    chunk = job_ids[i:i+WAVE]
    fstr = urllib.parse.quote(" or ".join(f"JobID eq {c}" for c in chunk))
    try:
        data = _get(f"/Activities?$filter={fstr}&$top={WAVE*20}")
        for a in data.get("value", []):
            if a.get("Cancelled"): continue
            jid   = a.get("JobID")
            atype = (a.get("ActivityType") or "").lower()
            start = (a.get("Start") or "")[:10]
            if jid and ("cm" in atype or "book" in atype or "install" in atype or "service" in atype):
                if jid not in booking_dates or start > booking_dates[jid]["date"]:
                    booking_dates[jid] = {"date": start, "type": a.get("ActivityType"), "status": a.get("Status")}
    except Exception as e:
        print(f"  Warning activities chunk {i}: {e}")
    if i % 300 == 0:
        print(f"  activities {min(i+WAVE,len(job_ids))}/{len(job_ids)}...", flush=True)

# 5. Fetch contacts for these jobs
print("Fetching contacts...")
contact_ids = list({j.get("ContactID") for j in service_jobs if j.get("ContactID")})
contacts = {}
for i in range(0, len(contact_ids), 20):
    chunk = contact_ids[i:i+20]
    fstr = urllib.parse.quote(" or ".join(f"ID eq {c}" for c in chunk))
    try:
        data = _get(f"/Contacts?$filter={fstr}&$top=40")
        for c in data.get("value", []):
            contacts[c["ID"]] = c
    except: pass

print(f"  {len(contacts)} contacts fetched\n")

# 6. Build output rows
rows = []
for j in service_jobs:
    jid    = j["ID"]
    ref    = j.get("Reference", "")
    status = j.get("Status", "")
    stage  = j.get("Stage", "")
    job_date   = (j.get("JobDate") or "")[:10]
    eta_date   = (j.get("ETADate") or "")[:10]
    close_date = (j.get("CloseDate") or "")[:10]
    cust_notes = j.get("CustomerNotes") or ""
    install_notes = j.get("InstallNotes") or ""

    # Contact details
    cid  = j.get("ContactID")
    cont = contacts.get(cid, {})
    cust_name = f"{cont.get('FirstName','')} {cont.get('LastName','')}".strip() or f"Contact {cid}"
    mobile    = cont.get("Mobile") or ""
    email_c   = cont.get("Email") or ""

    # Lines summary
    lines_for_job = job_lines.get(jid, [])
    products = ", ".join({l.get("Product","") for l in lines_for_job if l.get("Product")})
    line_stages = ", ".join({l.get("Stage","") for l in lines_for_job if l.get("Stage")})

    # Booking date from activities
    bk = booking_dates.get(jid, {})
    booking_date = bk.get("date", "")
    booking_type = bk.get("type", "")
    has_booking  = "Yes" if booking_date else "No"

    rows.append({
        "Job Reference":   ref,
        "Job ID":          jid,
        "Customer Name":   cust_name,
        "Mobile":          mobile,
        "Email":           email_c,
        "Job Date":        job_date,
        "Job Status":      status,
        "Job Stage":       stage,
        "Products":        products,
        "Line Stages":     line_stages,
        "ETA Date":        eta_date,
        "Has Booking":     has_booking,
        "Booking Date":    booking_date,
        "Booking Type":    booking_type,
        "Customer Notes":  cust_notes,
        "Install Notes":   install_notes,
    })

print(f"Total rows: {len(rows)}")
print(f"With booking: {sum(1 for r in rows if r['Has Booking']=='Yes')}")
print(f"Without booking: {sum(1 for r in rows if r['Has Booking']=='No')}")

# 7. Export to Excel
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Open Service Jobs"

COLS = list(rows[0].keys()) if rows else []

# Header style
hdr_fill = PatternFill("solid", fgColor="1E3A5F")
hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
body_font = Font(name="Arial", size=10)
green_fill = PatternFill("solid", fgColor="D1FAE5")
amber_fill = PatternFill("solid", fgColor="FEF3C7")
border = Border(bottom=Side(style="thin", color="E2E8F0"))

ws.append(COLS)
for cell in ws[1]:
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

for row in rows:
    ws.append([row.get(c, "") for c in COLS])
    row_idx = ws.max_row
    # Highlight booking status
    has_bk_col = COLS.index("Has Booking") + 1
    cell = ws.cell(row=row_idx, column=has_bk_col)
    if cell.value == "Yes":
        cell.fill = green_fill
        cell.font = Font(name="Arial", size=10, bold=True, color="065F46")
    else:
        cell.fill = amber_fill
        cell.font = Font(name="Arial", size=10, bold=True, color="92400E")
    for c in ws[row_idx]:
        c.font = body_font
        c.border = border
        c.alignment = Alignment(horizontal="left", vertical="center")

# Column widths
widths = {"Job Reference":14,"Job ID":10,"Customer Name":22,"Mobile":15,"Email":28,
          "Job Date":12,"Job Status":18,"Job Stage":18,"Products":30,"Line Stages":30,
          "ETA Date":12,"Has Booking":12,"Booking Date":14,"Booking Type":20,
          "Customer Notes":35,"Install Notes":35}
for i, col in enumerate(COLS, 1):
    ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 16)

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

out = "maria_rundle_open_services.xlsx"
wb.save(out)
print(f"\n>> Saved: {out}")
