import urllib.request, urllib.parse, json, base64, time, re
from collections import Counter, defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from insyte_env import EMAIL, KEY
BASE  = "https://api.myinsyte.com.au/v2"
AUTH  = "Basic " + base64.b64encode(f"{EMAIL}:{KEY}".encode()).decode()

def _get(path, retries=5):
    url = f"{BASE}/{path.lstrip('/')}"
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers={"Authorization": AUTH, "Accept": "application/json"})
            with urllib.request.urlopen(req, timeout=120) as r:
                return json.loads(r.read())
        except Exception as e:
            if attempt < retries - 1: time.sleep(2 ** attempt)
            else: raise

def fetch_all(path, label):
    PAGE, rows, skip, pg = 500, [], 0, 1
    sep = "&" if "?" in path else "?"
    while True:
        print(f"  {label} p{pg} ({len(rows)})...", flush=True)
        data = _get(f"{path}{sep}$top={PAGE}&$skip={skip}")
        page = data.get("value", data if isinstance(data, list) else [])
        rows.extend(page)
        if len(page) < PAGE: break
        skip += PAGE; pg += 1
    return rows

def base_ref(r):
    return re.sub(r'-\d+$', '', str(r or ''))

OPT_KEYS = [f"DisplayOption{i}" for i in range(1, 11)]

# ── 1. Fetch confirmed lines
print("Fetching 2026 confirmed lines...")
f = urllib.parse.quote("OrderDate ge 2026-01-01T00:00:00Z and OrderDate le 2026-06-29T23:59:59Z")
conf_lines = fetch_all(f"/JobLines?$filter={f}", "conf lines")
print(f"  {len(conf_lines):,} confirmed lines\n")

print("Fetching users...")
users = fetch_all("/Users", "users")
user_map = {u["ID"]: (u.get("FullName") or f"{u.get('FirstName','')} {u.get('LastName','')}".strip()) for u in users}

# ── 2. Fetch ALL 2026 jobs (to also get quote-stage job IDs)
print("Fetching all 2026 jobs...")
jf = urllib.parse.quote("JobDate ge 2026-01-01T00:00:00Z and JobDate le 2026-06-29T23:59:59Z")
all_jobs_2026 = fetch_all(f"/Jobs?$filter={jf}", "all jobs")
job_map = {j["ID"]: j for j in all_jobs_2026}
print(f"  {len(job_map):,} jobs\n")

# ── 3. Fetch lines for quote-stage jobs (brick fixing only appears there)
confirmed_job_ids = {l["JobID"] for l in conf_lines if l.get("JobID")}
quote_job_ids = [j["ID"] for j in all_jobs_2026
                 if j.get("Stage") == "stage_job_quote"
                 and j.get("JobType") != "type_job_service"
                 and j["ID"] not in confirmed_job_ids]
print(f"Fetching lines for {len(quote_job_ids):,} quote-stage jobs (chunk 15)...")
CHUNK = 15
quote_lines = []
for i in range(0, len(quote_job_ids), CHUNK):
    chunk = quote_job_ids[i:i+CHUNK]
    fstr  = urllib.parse.quote(" or ".join(f"JobID eq {c}" for c in chunk))
    try:
        data = _get(f"/JobLines?$filter={fstr}&$top={CHUNK*20}")
        quote_lines.extend(data.get("value", []))
    except Exception as e:
        print(f"  Warning chunk {i}: {e}")
    if i % 750 == 0:
        print(f"  {min(i+CHUNK, len(quote_job_ids)):,}/{len(quote_job_ids):,} jobs...", flush=True)
print(f"  {len(quote_lines):,} quote lines\n")

all_lines = conf_lines + quote_lines

# ── 4. Enumerate ALL DisplayOption key:value pairs
print("Enumerating all DisplayOption keys...")
key_vals = defaultdict(Counter)
for l in all_lines:
    for opt in OPT_KEYS:
        val = (l.get(opt) or "").strip()
        if not val: continue
        colon = val.find(":")
        if colon > 0:
            prefix = val[:colon].strip()
            value  = val[colon+1:].strip()
            key_vals[prefix][value] += 1

print("\n--- Fixing values (all stages) ---")
for v, c in key_vals.get("Fixing", Counter()).most_common():
    print(f"  {c:5d}  {v!r}")

print("\n--- Keys with 'take', 'down', 'existing', 'removal', 'replace' ---")
for key in sorted(key_vals):
    kl = key.lower()
    if any(t in kl for t in ["take", "down", "existing", "remov", "old blind", "replace"]):
        print(f"\n  {key!r} ({sum(key_vals[key].values())} total):")
        for v, c in key_vals[key].most_common():
            print(f"    {c:5d}  {v!r}")

fixing_key = "Fixing"
print(f"\nUsing: fixing_key={fixing_key!r}\n")

# helper
def get_opt(line, key):
    for opt in OPT_KEYS:
        val = (line.get(opt) or "").strip()
        if val.startswith(key + ":"):
            return val[len(key)+1:].strip()
    return None

EXCL_STATUS = {"status_job_line_cancelled"}
EXCL_TYPE   = {"type_job_line_remake", "type_job_line_service", "type_job_line_alter"}
# For brick: include all non-cancelled lines (quotes too); for takedown use confirmed only
active_all  = [l for l in all_lines
               if l.get("Status") not in EXCL_STATUS
               and l.get("LineType") not in EXCL_TYPE]
active_conf = [l for l in conf_lines
               if l.get("Status") not in EXCL_STATUS
               and l.get("LineType") not in EXCL_TYPE
               and l.get("Stage") != "stage_job_line_unconfirmed"]

# ── 5. BRICK FIXING
brick_lines = [l for l in active_all
               if "brick" in (get_opt(l, fixing_key) or "").lower()]

brick_jobs = defaultdict(list)
for l in brick_lines:
    job = job_map.get(l["JobID"], {})
    br  = base_ref(job.get("Reference") or str(l["JobID"]))
    brick_jobs[br].append(l)

print(f"BRICK FIXING: {len(brick_lines)} lines across {len(brick_jobs)} unique orders")

# ── 6. TAKEDOWN — using confirmed lines
td_lines = []
for l in active_conf:
    for opt in OPT_KEYS:
        val = (l.get(opt) or "").strip().lower()
        if ("take" in val and "down" in val) or "existing" in val or "takedown" in val:
            td_lines.append(l)
            break

td_product_lines = [l for l in conf_lines if "take down" in (l.get("Product") or "").lower()]
all_td = {l.get("ID"): l for l in td_lines + td_product_lines}
td_by_job = defaultdict(list)
for l in all_td.values():
    job = job_map.get(l["JobID"], {})
    br  = base_ref(job.get("Reference") or str(l["JobID"]))
    td_by_job[br].append(l)

# Check if roller blind product exists in same job
roller_td_jobs = {}
for br, ls in td_by_job.items():
    same_job_lines = [x for x in active_conf if job_map.get(x.get("JobID"),{}).get("Reference","").startswith(br)]
    has_roller = any("roller" in (x.get("Product") or "").lower() for x in same_job_lines + ls)
    roller_td_jobs[br] = {"lines": ls, "has_roller": has_roller}

roller_only = {br: v for br, v in roller_td_jobs.items() if v["has_roller"]}
print(f"TAKEDOWN JOBS: {len(td_by_job)} total, {len(roller_only)} with roller blind in same job")

# ── 7. Fetch contacts
all_relevant_job_ids = (
    {l["JobID"] for l in brick_lines} |
    {l["JobID"] for ls in td_by_job.values() for l in ls}
)
contact_ids = list({job_map[jid].get("ContactID") for jid in all_relevant_job_ids
                    if job_map.get(jid, {}).get("ContactID")})
contacts = {}
for i in range(0, len(contact_ids), 20):
    chunk = contact_ids[i:i+20]
    fstr  = urllib.parse.quote(" or ".join(f"ID eq {c}" for c in chunk))
    try:
        data = _get(f"/Contacts?$filter={fstr}&$top=40")
        for c in data.get("value", []):
            contacts[c["ID"]] = c
    except: pass
print(f"  {len(contacts)} contacts fetched\n")

def contact_info(jid):
    cid  = job_map.get(jid, {}).get("ContactID")
    c    = contacts.get(cid, {})
    name = f"{c.get('FirstName','')} {c.get('LastName','')}".strip() or f"Contact {cid}"
    return name, c.get("Mobile",""), c.get("Email","")

def rep_name(jid):
    return user_map.get(job_map.get(jid,{}).get("SalesRepID"), "")

def fmt_date(d):
    return (d or "")[:10]

# ── 8. Build Excel
wb  = openpyxl.Workbook()
NAV = "1E3A5F"; WHT = "FFFFFF"; GRN = "D1FAE5"; AMB = "FEF3C7"

def hdr_style(cell):
    cell.font      = Font(name="Arial", bold=True, color=WHT, size=10)
    cell.fill      = PatternFill("solid", fgColor=NAV)
    cell.alignment = Alignment(horizontal="left", vertical="center")

def body_style(cell):
    cell.font      = Font(name="Arial", size=10)
    cell.border    = Border(bottom=Side(style="thin", color="E2E8F0"))
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

def write_sheet(ws, headers, rows_data, col_widths, tab_color=None):
    if tab_color: ws.sheet_properties.tabColor = tab_color
    ws.append(headers)
    for cell in ws[1]: hdr_style(cell)
    for row in rows_data:
        ws.append(row)
        for cell in ws[ws.max_row]: body_style(cell)
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# Sheet 1: Brick Fixings
ws1 = wb.active
ws1.title = "Brick Fixings"
hdrs1 = ["Job Reference","Base Ref","Job Date","Order Date","Customer Name","Mobile","Email",
         "Sales Rep","Product","Fixing Value","Job Stage","Line Stage","Sale Price Ex Tax"]
rows1 = []
for l in sorted(brick_lines, key=lambda x: x.get("OrderDate") or x.get("JobDate") or ""):
    job  = job_map.get(l["JobID"], {})
    br   = base_ref(job.get("Reference") or str(l["JobID"]))
    name, mob, eml = contact_info(l["JobID"])
    rows1.append([
        job.get("Reference",""), br,
        fmt_date(job.get("JobDate")),
        fmt_date(l.get("OrderDate")),
        name, mob, eml,
        rep_name(l["JobID"]),
        l.get("Product",""),
        get_opt(l, fixing_key) or "",
        (job.get("Stage") or "").replace("stage_job_",""),
        (l.get("Stage") or "").replace("stage_job_line_",""),
        round(l.get("DiscountedPriceExTax") or 0, 2),
    ])
write_sheet(ws1, hdrs1, rows1, [16,12,12,12,22,15,28,20,28,14,14,16,14], "1E3A5F")

# Sheet 2: All Takedowns
ws2 = wb.create_sheet("Takedowns - All")
hdrs2 = ["Job Reference","Base Ref","Order Date","Customer Name","Mobile","Email",
         "Sales Rep","Product","Takedown/Existing Option","Has Roller in Job",
         "Line Stage","Sale Price Ex Tax"]
rows2 = []
roller_col_idx = hdrs2.index("Has Roller in Job") + 1
for br, ls in sorted(td_by_job.items()):
    for l in ls:
        job  = job_map.get(l["JobID"], {})
        name, mob, eml = contact_info(l["JobID"])
        td_val = ""
        for opt in OPT_KEYS:
            val = (l.get(opt) or "").strip()
            vl  = val.lower()
            if ("take" in vl and "down" in vl) or "existing" in vl or "takedown" in vl:
                td_val = val; break
        if not td_val and "take down" in (l.get("Product") or "").lower():
            td_val = f"Product: {l.get('Product','')}"
        has_roller = "Yes" if roller_td_jobs.get(br, {}).get("has_roller") else "No"
        rows2.append([
            job.get("Reference",""), br,
            fmt_date(l.get("OrderDate") or job.get("JobDate")),
            name, mob, eml,
            rep_name(l["JobID"]),
            l.get("Product",""),
            td_val, has_roller,
            (l.get("Stage") or "").replace("stage_job_line_",""),
            round(l.get("DiscountedPriceExTax") or 0, 2),
        ])
write_sheet(ws2, hdrs2, rows2, [16,12,12,22,15,28,20,28,30,12,16,14], "7C3AED")
for ri in range(2, ws2.max_row + 1):
    cell = ws2.cell(row=ri, column=roller_col_idx)
    cell.fill = PatternFill("solid", fgColor=GRN if cell.value == "Yes" else AMB)
    cell.font = Font(name="Arial", size=10, bold=True,
                     color="065F46" if cell.value == "Yes" else "92400E")

# Sheet 3: Roller Blind Takedowns only
ws3 = wb.create_sheet("Roller Takedowns")
rows3 = [r for r in rows2 if r[9] == "Yes"]
write_sheet(ws3, hdrs2, rows3, [16,12,12,22,15,28,20,28,30,12,16,14], "059669")
for ri in range(2, ws3.max_row + 1):
    cell = ws3.cell(row=ri, column=roller_col_idx)
    cell.fill = PatternFill("solid", fgColor=GRN)
    cell.font = Font(name="Arial", size=10, bold=True, color="065F46")

# Sheet 4: Summary
ws4 = wb.create_sheet("Summary")
ws4.sheet_properties.tabColor = "D97706"
summary_rows = [
    ["Metric", "Count"],
    ["", ""],
    ["BRICK FIXINGS", ""],
    ["Lines with Fixing = Brick", len(brick_lines)],
    ["Unique orders (base ref) with Brick fixing", len(brick_jobs)],
    ["", ""],
    ["TAKEDOWNS (confirmed orders only)", ""],
    ["Total jobs with a takedown/existing option", len(td_by_job)],
    ["  — of which include a Roller Blind product", len(roller_only)],
    ["  — takedown jobs without a Roller Blind", len(td_by_job) - len(roller_only)],
    ["", ""],
    ["DATA SCOPE", ""],
    ["Period", "2026 YTD (Jan 1 – Jun 29)"],
    ["Brick source", "All lines incl. quote-stage"],
    ["Takedown source", "Confirmed order lines only"],
]
for row in summary_rows:
    ws4.append(row)
ws4["A1"].font = Font(name="Arial", bold=True, size=10, color=WHT)
ws4["B1"].font = Font(name="Arial", bold=True, size=10, color=WHT)
ws4["A1"].fill = PatternFill("solid", fgColor=NAV)
ws4["B1"].fill = PatternFill("solid", fgColor=NAV)
for ri in range(2, ws4.max_row + 1):
    for cell in ws4[ri]:
        cell.font = Font(name="Arial", size=10)
        cell.alignment = Alignment(horizontal="left")
    label = ws4.cell(row=ri, column=1).value or ""
    if label and not label.startswith(" "):
        for cell in ws4[ri]:
            cell.font = Font(name="Arial", bold=True, size=10)
ws4.column_dimensions["A"].width = 48
ws4.column_dimensions["B"].width = 20

out = "brick_and_takedown_2026.xlsx"
wb.save(out)
print(f">> Saved: {out}")
print(f"   Sheet 1 Brick Fixings:      {len(rows1)} lines / {len(brick_jobs)} orders")
print(f"   Sheet 2 All Takedowns:      {len(rows2)} rows")
print(f"   Sheet 3 Roller Takedowns:   {len(rows3)} rows")
