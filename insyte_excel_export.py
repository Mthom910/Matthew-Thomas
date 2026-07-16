"""
insyte_excel_export.py
Pulls all available data from the Insyte OData API and generates a
comprehensive Excel workbook. One row per unique base reference number,
with columns from every API endpoint joined together.

Run:  python insyte_excel_export.py
Out:  insyte_export_YYYYMMDD_HHMM.xlsx
"""

import base64, json, urllib.request, urllib.parse, urllib.error
import re, os, sys
from datetime import datetime, date
from collections import defaultdict

# -- Try pandas + openpyxl; install if missing
try:
    import pandas as pd
except ImportError:
    os.system(f"{sys.executable} -m pip install pandas openpyxl")
    import pandas as pd

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# CONFIG
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

API_BASE  = "https://api.myinsyte.com.au/v2"
from insyte_env import EMAIL as API_EMAIL, KEY as API_KEY

# Date range for all time-filtered endpoints (JobLines, Activities, Invoices)
# Set LINES_TO = "" to fetch from LINES_FROM to today (open-ended)
LINES_FROM = "2026-04-01"
LINES_TO   = "2026-05-31"

# Stage sets — adjust if your Insyte instance uses different identifiers
WON_STAGES = {
    'stage_job_line_confirmed','stage_job_line_confirmed_cm',
    'stage_job_line_cm_booked','stage_job_line_cm_inprogress',
    'stage_job_line_cm_completed','stage_job_line_cm_completed_confirmed',
    'stage_job_line_cm_completed_reschedule','stage_job_line_cm_completed_unconfirmed',
    'stage_job_line_supply','stage_job_line_supply_failed',
    'stage_job_line_dispatch','stage_job_line_invoicing','stage_job_line_restocked',
}
LOST_STAGES = {
    'stage_job_line_cancelled','status_invoice_line_cancelled',
    'stage_job_line_expired','stage_job_line_lost',
}

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# API HELPERS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _auth():
    token = base64.b64encode(f"{API_EMAIL}:{API_KEY}".encode()).decode()
    return {"Authorization": f"Basic {token}", "Accept": "application/json"}

def _get(path, retries=5):
    """GET with retry/backoff on timeout or transient errors."""
    import time, http.client
    url = f"{API_BASE}/{path.lstrip('/')}"
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers=_auth())
            with urllib.request.urlopen(req, timeout=120) as r:
                return json.loads(r.read())
        except (TimeoutError, ConnectionResetError, ConnectionError,
                urllib.error.URLError, http.client.RemoteDisconnected,
                http.client.IncompleteRead) as e:
            wait = 3 ** attempt  # 1s, 3s, 9s, 27s, 81s
            if attempt < retries - 1:
                print(f"  [retry {attempt+1}/{retries-1} in {wait}s] {type(e).__name__}: {e}")
                time.sleep(wait)
            else:
                raise

def fetch_all(path, label=""):
    """Paginate through an OData endpoint using $skip (Insyte does not return nextLink)."""
    PAGE = 500
    clean = re.sub(r'([&?])\$top=\d+', r'\1', path)
    clean = re.sub(r'([&?])\$skip=\d+', r'\1', clean).rstrip('?&')
    sep   = '&' if '?' in clean else '?'
    rows, skip = [], 0
    while True:
        url = f"{clean}{sep}$top={PAGE}&$skip={skip}"
        data = _get(url)
        page = data.get('value', data) if isinstance(data, dict) else data
        rows.extend(page)
        print(f"  {label or path.split('?')[0]}: +{len(page)} (total {len(rows)})")
        if len(page) < PAGE:
            break
        skip += PAGE
    return rows

def batch_jobs(job_ids):
    """Fetch Jobs by ID list in chunks of 15 with 5-way concurrency (sequential fallback)."""
    import concurrent.futures
    CHUNK = 15
    chunks = [job_ids[i:i+CHUNK] for i in range(0, len(job_ids), CHUNK)]
    job_map = {}
    def fetch_chunk(ids):
        f = urllib.parse.quote(" or ".join(f"ID eq {i}" for i in ids))
        return fetch_all(f"/Jobs?$expand=Address&$filter={f}", label="  /Jobs chunk")
    with concurrent.futures.ThreadPoolExecutor(max_workers=5) as ex:
        for result in ex.map(fetch_chunk, chunks):
            for j in result:
                job_map[j['ID']] = j
    return job_map

def base_ref(ref):
    """Strip trailing variation suffix: 'J0012345-2' → 'J0012345'."""
    return re.sub(r'-\d+$', '', str(ref)) if ref else ''

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# DATA FETCHING
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def _date_filter(field, from_dt, to_dt=""):
    """Build OData date range filter string (unencoded)."""
    parts = [f"{field} ge {from_dt}T00:00:00Z"]
    if to_dt:
        parts.append(f"{field} le {to_dt}T23:59:59Z")
    return " and ".join(parts)

def fetch_all_data():
    print("\n-- Fetching data from Insyte API --")

    line_filter = urllib.parse.quote(_date_filter("OrderDate", LINES_FROM, LINES_TO))
    print("\nJobLines (ordered, from %s%s):" % (LINES_FROM, f" to {LINES_TO}" if LINES_TO else "+"))
    lines = fetch_all(f"/JobLines?$filter={line_filter}", "JobLines")

    print("\nUsers:")
    users = fetch_all("/Users", "Users")

    print("\nOpportunities:")
    opps = fetch_all("/Opportunities", "Opportunities")

    act_filter = urllib.parse.quote(_date_filter("Start", LINES_FROM, LINES_TO))
    print("\nActivities (from %s%s):" % (LINES_FROM, f" to {LINES_TO}" if LINES_TO else "+"))
    acts = fetch_all(f"/Activities?$filter={act_filter}", "Activities")

    inv_filter = urllib.parse.quote(_date_filter("InvoiceDate", LINES_FROM, LINES_TO))
    print("\nInvoices (from %s%s):" % (LINES_FROM, f" to {LINES_TO}" if LINES_TO else "+"))
    invoices = fetch_all(f"/Invoices?$filter={inv_filter}", "Invoices")

    print("\nJobs (batch by ID from lines):")
    job_ids = list({l['JobID'] for l in lines if l.get('JobID')})
    print(f"  {len(job_ids)} unique job IDs to fetch")
    job_map = batch_jobs(job_ids)
    jobs = list(job_map.values())

    print(f"\n[OK] Loaded: {len(lines)} lines · {len(jobs)} jobs · {len(opps)} opportunities "
          f"· {len(acts)} activities · {len(invoices)} invoices · {len(users)} users\n")

    return lines, jobs, opps, acts, invoices, users

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# BUILD DATAFRAMES
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def build_dataframes(lines, jobs, opps, acts, invoices, users):
    user_map = {u['ID']: u for u in users}

    def rep_name(uid):
        u = user_map.get(uid, {})
        return u.get('FullName') or f"{u.get('FirstName','')} {u.get('LastName','')}".strip() or f"Rep {uid}"

    # -- Job lookup
    job_map = {j['ID']: j for j in jobs}

    # -- JobLines DataFrame
    line_rows = []
    for l in lines:
        job  = job_map.get(l.get('JobID'), {})
        addr = job.get('Address') or {}
        std  = l.get('StandardPriceExTax')  or 0
        disc = l.get('DiscountedPriceExTax') or std
        price_inc = l.get('DiscountedPriceIncTax') or l.get('StandardPriceIncTax') or disc
        cost      = l.get('DiscountedCostExTax')   or l.get('StandardCostExTax')   or 0
        inst      = l.get('StandardIntallCostExTax')    or 0
        deliv     = l.get('StandardDeliveryCostExTax')  or 0
        gp        = price_inc - cost - inst - deliv
        disc_pct  = round((1 - disc / std) * 100, 2) if std > 0 else 0
        sqm       = round(((l.get('Width') or 0) * (l.get('Drop') or 0)) / 1_000_000, 4)
        stage     = (l.get('Stage') or '').lower()
        raw_ref   = job.get('Reference') or str(l.get('JobID',''))
        br        = base_ref(raw_ref)

        line_rows.append({
            'Base Reference':         br,
            'Job Reference':          raw_ref,
            'Job ID':                 l.get('JobID'),
            'Job Line No':            l.get('JobLineNo'),
            'Order Date':             l.get('OrderDate','')[:10] if l.get('OrderDate') else '',
            'Job Date':               job.get('JobDate','')[:10] if job.get('JobDate') else '',
            'Sales Rep':              rep_name(job.get('SalesRepID')),
            'Sales Rep ID':           job.get('SalesRepID'),
            'State':                  addr.get('State',''),
            'Suburb':                 addr.get('Suburb',''),
            'Postcode':               addr.get('PostCode',''),
            'Customer Name':          job.get('CustomerName') or job.get('ContactName') or '',
            'Stage':                  l.get('Stage',''),
            'Stage Classification':   'Won' if stage in WON_STAGES else 'Lost/Cancelled' if stage in LOST_STAGES else 'Quote/Open',
            'Product':                l.get('Product') or l.get('ProductCode') or '',
            'Product Group':          l.get('MultiProduct') or '',
            'Supplier':               l.get('Supplier') or '',
            'Width (mm)':             l.get('Width') or 0,
            'Drop (mm)':              l.get('Drop') or 0,
            'SQM':                    sqm,
            'Price Std Ex Tax':       round(std, 2),
            'Price Disc Ex Tax':      round(disc, 2),
            'Price Inc Tax':          round(price_inc, 2),
            'Cost Ex Tax':            round(cost, 2),
            'Install Cost':           round(inst, 2),
            'Delivery Cost':          round(deliv, 2),
            'Total Cost':             round(cost + inst + deliv, 2),
            'GP $':                   round(gp, 2),
            'GP %':                   round(gp / price_inc * 100, 2) if price_inc > 0 else 0,
            'Discount %':             disc_pct,
        })
    df_lines = pd.DataFrame(line_rows)

    # -- Opportunities DataFrame
    opp_rows = []
    for o in opps:
        raw_ref = o.get('Reference') or o.get('JobReference') or str(o.get('ID',''))
        br      = base_ref(raw_ref)
        opp_rows.append({
            'Base Reference':       br,
            'Opp Reference':        raw_ref,
            'Opp ID':               o.get('ID'),
            'Opp Status':           o.get('Status',''),
            'Pipeline Stage':       o.get('PipelineStage') or o.get('Stage') or '',
            'Expected Revenue':     o.get('ExpectedRevenue') or o.get('Value') or 0,
            'Opp Rep':              rep_name(o.get('RepresentativeID')),
            'Opp Rep ID':           o.get('RepresentativeID'),
            'Opp Created':          (o.get('CreatedOn') or o.get('CreatedDate') or '')[:10],
            'Opp Closed':           (o.get('ClosedOn') or o.get('ClosedDate') or '')[:10],
            'Outcome Reason':       o.get('OutcomeReason') or o.get('LostReason') or '',
            'Lead Source':          o.get('LeadSource') or '',
            'Opp Notes':            o.get('Notes') or o.get('Description') or '',
        })
    df_opps = pd.DataFrame(opp_rows) if opp_rows else pd.DataFrame(columns=['Base Reference'])

    # -- Activities grouped by base reference (via job reference match)
    # Activities link to contacts/jobs via RepresentativeID + date proximity
    # We summarise by rep as a best-effort grouping
    act_rows = []
    for a in acts:
        act_rows.append({
            'Activity ID':      a.get('ID'),
            'Activity Type':    a.get('ActivityType') or a.get('Type') or '',
            'Activity Status':  a.get('Status') or '',
            'Start':            (a.get('Start') or '')[:16],
            'End':              (a.get('End') or '')[:16],
            'Rep ID':           a.get('RepresentativeID'),
            'Rep':              rep_name(a.get('RepresentativeID')),
            'Subject':          a.get('Subject') or a.get('Title') or '',
            'Closed':           a.get('Closed'),
            'Contact Name':     a.get('ContactName') or '',
            'Contact Phone':    a.get('ContactPhone') or '',
            'Contact Email':    a.get('ContactEmail') or '',
            'Notes':            a.get('Notes') or '',
        })
    df_acts = pd.DataFrame(act_rows) if act_rows else pd.DataFrame()

    # -- Invoices grouped by job reference
    inv_rows = []
    for inv in invoices:
        raw_ref = inv.get('Reference') or inv.get('JobReference') or str(inv.get('ID',''))
        br      = base_ref(raw_ref)
        inv_rows.append({
            'Base Reference':       br,
            'Invoice Reference':    raw_ref,
            'Invoice ID':           inv.get('ID'),
            'Invoice Status':       inv.get('Status') or '',
            'Invoice Date':         (inv.get('InvoiceDate') or inv.get('Date') or '')[:10],
            'Due Date':             (inv.get('DueDate') or '')[:10],
            'Invoice Amount':       inv.get('Amount') or inv.get('TotalAmount') or 0,
            'Amount Paid':          inv.get('AmountPaid') or inv.get('PaidAmount') or 0,
            'Outstanding':          (inv.get('Amount') or 0) - (inv.get('AmountPaid') or 0),
            'Credit Note':          inv.get('CreditNote') or False,
        })
    df_invs = pd.DataFrame(inv_rows) if inv_rows else pd.DataFrame(columns=['Base Reference'])

    # -- Users DataFrame
    user_rows = []
    for u in users:
        user_rows.append({
            'User ID':          u.get('ID'),
            'Full Name':        u.get('FullName') or f"{u.get('FirstName','')} {u.get('LastName','')}".strip(),
            'First Name':       u.get('FirstName',''),
            'Last Name':        u.get('LastName',''),
            'Email':            u.get('Email',''),
            'Phone':            u.get('Phone') or u.get('Mobile') or '',
            'Role':             u.get('Role') or u.get('UserRole') or '',
            'Active':           u.get('IsActive') if 'IsActive' in u else u.get('Active',''),
            'Business Unit':    u.get('BusinessUnit') or u.get('BusinessUnitName') or '',
        })
    df_users = pd.DataFrame(user_rows)

    return df_lines, df_opps, df_acts, df_invs, df_users

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# BUILD SUMMARY (one row per unique base reference)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

def build_summary(df_lines, df_opps, df_invs):
    if df_lines.empty:
        return pd.DataFrame()

    def join_unique(series, sep=', '):
        return sep.join(sorted({str(v) for v in series.dropna() if str(v).strip()}))

    # -- Aggregate job lines by base reference
    grp = df_lines.groupby('Base Reference')
    summary = pd.DataFrame({
        'All Job References':      grp['Job Reference'].apply(join_unique),
        'Job IDs':                 grp['Job ID'].apply(lambda s: join_unique(s.astype(str))),
        'Sales Rep':               grp['Sales Rep'].apply(lambda s: join_unique(s)),
        'State':                   grp['State'].apply(lambda s: join_unique(s)),
        'Suburb':                  grp['Suburb'].apply(lambda s: join_unique(s)),
        'Postcode':                grp['Postcode'].apply(lambda s: join_unique(s)),
        'Customer Name':           grp['Customer Name'].apply(lambda s: join_unique(s)),
        'First Order Date':        grp['Order Date'].min(),
        'Last Order Date':         grp['Order Date'].max(),
        'Job Date':                grp['Job Date'].min(),
        '# Lines':                 grp.size(),
        'Products':                grp['Product'].apply(join_unique),
        'Product Groups':          grp['Product Group'].apply(join_unique),
        'Primary Product':         grp.apply(lambda g: g.loc[g['Price Inc Tax'].idxmax(), 'Product']),
        'Supplier(s)':             grp['Supplier'].apply(join_unique),
        'Total Value (inc Tax)':   grp['Price Inc Tax'].sum().round(2),
        'Total Cost':              grp['Total Cost'].sum().round(2),
        'GP $':                    grp['GP $'].sum().round(2),
        'Avg Discount %':          grp['Discount %'].mean().round(2),
        'Total SQM':               grp['SQM'].sum().round(4),
        'Has Won Lines':           grp['Stage Classification'].apply(lambda s: 'Yes' if 'Won' in s.values else 'No'),
        'Has Open Lines':          grp['Stage Classification'].apply(lambda s: 'Yes' if 'Quote/Open' in s.values else 'No'),
        'Has Lost Lines':          grp['Stage Classification'].apply(lambda s: 'Yes' if 'Lost/Cancelled' in s.values else 'No'),
        'All Stages':              grp['Stage'].apply(join_unique),
    }).reset_index()

    summary['GP %'] = (summary['GP $'] / summary['Total Value (inc Tax)'] * 100).round(2)

    # -- Join Opportunities (one opp per base ref — take most recent if multiple)
    if not df_opps.empty and 'Base Reference' in df_opps.columns:
        opp_agg = df_opps.sort_values('Opp Created', ascending=False).groupby('Base Reference').agg(
            Opp_Status      = ('Opp Status',       'first'),
            Pipeline_Stage  = ('Pipeline Stage',   'first'),
            Expected_Revenue= ('Expected Revenue', 'sum'),
            Opp_Rep         = ('Opp Rep',          'first'),
            Opp_Created     = ('Opp Created',      'min'),
            Opp_Closed      = ('Opp Closed',       'max'),
            Outcome_Reason  = ('Outcome Reason',   'first'),
            Lead_Source     = ('Lead Source',      'first'),
        ).reset_index()
        opp_agg.columns = ['Base Reference','Opp Status','Pipeline Stage','Expected Revenue',
                           'Opp Rep','Opp Created Date','Opp Closed Date','Outcome Reason','Lead Source']
        summary = summary.merge(opp_agg, on='Base Reference', how='left')

    # -- Join Invoice summary
    if not df_invs.empty and 'Base Reference' in df_invs.columns:
        inv_agg = df_invs.groupby('Base Reference').agg(
            Invoice_Status    = ('Invoice Status',  lambda s: join_unique(s)),
            Invoice_Total     = ('Invoice Amount',  'sum'),
            Amount_Paid       = ('Amount Paid',     'sum'),
            Outstanding       = ('Outstanding',     'sum'),
        ).reset_index()
        inv_agg.columns = ['Base Reference','Invoice Status','Invoice Total','Amount Paid','Outstanding']
        summary = summary.merge(inv_agg, on='Base Reference', how='left')

    return summary

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# EXCEL WRITING
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

# Colour palette
C_DARK_BLUE  = "1E3A5F"
C_MID_BLUE   = "2E6DB4"
C_LIGHT_BLUE = "D6E4F0"
C_GREEN      = "1A7A4A"
C_LIGHT_GREEN= "D5F5E3"
C_AMBER      = "E67E22"
C_LIGHT_AMB  = "FDEBD0"
C_RED        = "C0392B"
C_LIGHT_RED  = "FADBD8"
C_GREY_HDR   = "2C3E50"
C_GREY_LIGHT = "ECF0F1"
C_WHITE      = "FFFFFF"
C_STRIPE     = "F8FAFB"

def _hdr_style(col, fill_hex=C_DARK_BLUE):
    col.font      = Font(name='Arial', bold=True, color=C_WHITE, size=9)
    col.fill      = PatternFill('solid', fgColor=fill_hex)
    col.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def _border(thick=False):
    s = 'medium' if thick else 'thin'
    side = Side(style=s, color='BBBBBB')
    return Border(left=side, right=side, top=side, bottom=side)

def write_sheet(wb, title, df, header_color=C_DARK_BLUE, freeze='A2',
                pct_cols=None, money_cols=None, num_cols=None):
    """Write a DataFrame to a named sheet with full formatting."""
    ws = wb.create_sheet(title)
    if df.empty:
        ws.append(['No data loaded'])
        return ws

    pct_cols   = pct_cols   or []
    money_cols = money_cols or []
    num_cols   = num_cols   or []

    # Header row
    for ci, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        _hdr_style(cell, header_color)
        cell.border = _border(thick=True)

    # Data rows
    for ri, row in enumerate(df.itertuples(index=False), 2):
        fill = PatternFill('solid', fgColor=C_WHITE if ri % 2 == 0 else C_STRIPE)
        for ci, val in enumerate(row, 1):
            col_name = df.columns[ci - 1]
            # Convert NaN/None to empty
            if val is None or (isinstance(val, float) and pd.isna(val)):
                val = ''
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = Font(name='Arial', size=9)
            cell.fill      = fill
            cell.border    = _border()
            cell.alignment = Alignment(vertical='center')
            # Number formats
            if col_name in money_cols and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif col_name in pct_cols and isinstance(val, (int, float)):
                cell.number_format = '0.00"%"'
                cell.alignment = Alignment(horizontal='right', vertical='center')
            elif col_name in num_cols and isinstance(val, (int, float)):
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right', vertical='center')

    # Auto column widths (capped)
    for ci, col_name in enumerate(df.columns, 1):
        col_letter = get_column_letter(ci)
        max_len = max(
            len(str(col_name)),
            df[col_name].astype(str).str.len().max() if not df[col_name].empty else 0
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 45)

    # Freeze pane
    ws.freeze_panes = freeze

    # Autofilter on header row
    ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}1"

    return ws

def add_status_colouring(ws, df, status_col_name, row_start=2):
    """Colour-code rows in a sheet based on a status column value."""
    if status_col_name not in df.columns:
        return
    sci = list(df.columns).index(status_col_name) + 1  # 1-based col index
    status_map = {
        'won':   C_LIGHT_GREEN,
        'open':  C_LIGHT_BLUE,
        'lost':  C_LIGHT_RED,
        'yes':   C_LIGHT_GREEN,
        'no':    C_LIGHT_RED,
        'paid':  C_LIGHT_GREEN,
        'outstanding': C_LIGHT_AMB,
    }
    for ri in range(row_start, ws.max_row + 1):
        val = (ws.cell(ri, sci).value or '').lower()
        for key, color in status_map.items():
            if key in val:
                ws.cell(ri, sci).fill = PatternFill('solid', fgColor=color)
                ws.cell(ri, sci).font = Font(name='Arial', size=9, bold=True,
                    color=C_GREEN if 'green' in color.lower() or color == C_LIGHT_GREEN
                          else C_RED if color == C_LIGHT_RED else C_AMBER)
                break

def build_workbook(summary, df_lines, df_opps, df_acts, df_invs, df_users):
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # -- SHEET 1: SUMMARY (one row per unique base reference)
    print("Writing Summary sheet…")
    money = ['Total Value (inc Tax)','Total Cost','GP $','Expected Revenue',
             'Invoice Total','Amount Paid','Outstanding']
    pct   = ['Avg Discount %','GP %']
    num   = ['# Lines','Total SQM']
    ws_sum = write_sheet(wb, 'Summary by Reference', summary,
                         header_color=C_DARK_BLUE, freeze='B2',
                         money_cols=money, pct_cols=pct, num_cols=num)
    if not summary.empty:
        add_status_colouring(ws_sum, summary, 'Has Won Lines')
        if 'Opp Status' in summary.columns:
            add_status_colouring(ws_sum, summary, 'Opp Status')
        if 'Invoice Status' in summary.columns:
            add_status_colouring(ws_sum, summary, 'Invoice Status')

    # -- SHEET 2: JOB LINES (all raw lines)
    print("Writing Job Lines sheet…")
    line_money = ['Price Std Ex Tax','Price Disc Ex Tax','Price Inc Tax',
                  'Cost Ex Tax','Install Cost','Delivery Cost','Total Cost','GP $']
    line_pct   = ['GP %','Discount %']
    line_num   = ['Width (mm)','Drop (mm)']
    write_sheet(wb, 'Job Lines', df_lines,
                header_color=C_MID_BLUE, freeze='C2',
                money_cols=line_money, pct_cols=line_pct, num_cols=line_num)

    # -- SHEET 3: OPPORTUNITIES
    print("Writing Opportunities sheet…")
    opp_money = ['Expected Revenue']
    ws_opp = write_sheet(wb, 'Opportunities', df_opps,
                         header_color=C_GREY_HDR, freeze='B2',
                         money_cols=opp_money)
    if not df_opps.empty:
        add_status_colouring(ws_opp, df_opps, 'Opp Status')

    # -- SHEET 4: ACTIVITIES / APPOINTMENTS
    print("Writing Activities sheet…")
    write_sheet(wb, 'Activities', df_acts,
                header_color="4A235A", freeze='B2')

    # -- SHEET 5: INVOICES
    print("Writing Invoices sheet…")
    inv_money = ['Invoice Amount','Amount Paid','Outstanding']
    ws_inv = write_sheet(wb, 'Invoices', df_invs,
                         header_color=C_GREEN, freeze='B2',
                         money_cols=inv_money)
    if not df_invs.empty:
        add_status_colouring(ws_inv, df_invs, 'Invoice Status')

    # -- SHEET 6: USERS / REPS
    print("Writing Users sheet…")
    write_sheet(wb, 'Users', df_users,
                header_color=C_GREY_HDR, freeze='B2')

    # -- SHEET 7: DATA NOTES
    ws_notes = wb.create_sheet('Data Notes')
    notes = [
        ['Insyte API Data Export'],
        ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M')],
        ['API Base:', API_BASE],
        ['Lines from:', LINES_FROM],
        [],
        ['Sheet', 'Description'],
        ['Summary by Reference', 'One row per unique base quote reference number. All data joined.'],
        ['Job Lines', 'Raw JobLines from Insyte — all line items for ordered jobs.'],
        ['Opportunities', 'All opportunities across all pipeline stages (Open/Won/Lost).'],
        ['Activities', 'All activity records (appointments, calls, follow-ups).'],
        ['Invoices', 'All invoice records and payment status.'],
        ['Users', 'All Insyte users and their details.'],
        [],
        ['Reference Logic', ''],
        ['Base Reference', 'The root quote number, e.g. J0012345'],
        ['Variation refs', 'J0012345-2, J0012345-3 etc. are grouped under J0012345'],
        [],
        ['Stage Classification', ''],
        ['Won', 'Confirmed, dispatched, invoiced, installed, etc.'],
        ['Quote/Open', 'Unconfirmed, in negotiation, awaiting decision'],
        ['Lost/Cancelled', 'Cancelled or expired lines'],
    ]
    for ri, row in enumerate(notes, 1):
        for ci, val in enumerate(row, 1):
            cell = ws_notes.cell(ri, ci, val)
            if ri == 1:
                cell.font = Font(name='Arial', bold=True, size=14, color=C_DARK_BLUE)
            elif ri == 6 or ci == 1 and ri > 5 and val:
                cell.font = Font(name='Arial', bold=True, size=9)
            else:
                cell.font = Font(name='Arial', size=9)
    ws_notes.column_dimensions['A'].width = 28
    ws_notes.column_dimensions['B'].width = 65

    return wb

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# MAIN
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

if __name__ == '__main__':
    start = datetime.now()

    # 1. Fetch
    lines, jobs, opps, acts, invoices, users = fetch_all_data()

    # 2. Build DataFrames
    print("Building data frames…")
    df_lines, df_opps, df_acts, df_invs, df_users = build_dataframes(
        lines, jobs, opps, acts, invoices, users
    )

    # 3. Build summary
    print("Building summary by reference…")
    summary = build_summary(df_lines, df_opps, df_invs)
    print(f"  {len(summary)} unique base references in summary")

    # 4. Write Excel
    print("Writing Excel workbook…")
    wb = build_workbook(summary, df_lines, df_opps, df_acts, df_invs, df_users)

    # 5. Save
    ts       = datetime.now().strftime('%Y%m%d_%H%M')
    year_tag = LINES_FROM[:7] + (f"_to_{LINES_TO[:7]}" if LINES_TO else "_onwards")
    out_path = os.path.join(os.path.dirname(__file__), f'insyte_export_{year_tag}_{ts}.xlsx')
    wb.save(out_path)

    elapsed = (datetime.now() - start).seconds
    print(f"\n[DONE] Saved: {out_path}")
    print(f"   {len(summary)} reference rows · {len(df_lines)} job lines · "
          f"{len(df_opps)} opportunities · {len(df_acts)} activities")
    print(f"   Completed in {elapsed}s")
