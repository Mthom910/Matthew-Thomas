"""
Export the Product Sales Analysis (section 2 of insyte_consultant_product_dashboard.html)
to a multi-sheet Excel workbook, for the current calendar year to date.

Methodology matches the dashboard exactly:
  - Core revenue excludes cancelled, unconfirmed-stage, and remake lines.
  - Service-type lines/jobs stay in the core total but are flagged separately.
  - Unconfirmed & remake lines are tracked in their own sheet, not discarded.
"""
import urllib.request, urllib.parse, json, base64, time, re
from collections import defaultdict
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from insyte_env import EMAIL, KEY
BASE  = "https://api.myinsyte.com.au/v2"
AUTH  = "Basic " + base64.b64encode(f"{EMAIL}:{KEY}".encode()).decode()

TODAY = date.today()
YEAR_FROM = f"{TODAY.year}-01-01"
YEAR_TO   = TODAY.isoformat()
OUT_FILE  = f"Product_Sales_Analysis_{TODAY.year}.xlsx"

# ── HTTP helpers ──────────────────────────────────────────────
def _get(path, retries=5):
    url = f"{BASE}/{path.lstrip('/')}"
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers={"Authorization": AUTH, "Accept": "application/json"})
            with urllib.request.urlopen(req, timeout=120) as r:
                return json.loads(r.read())
        except Exception as e:
            if attempt < retries - 1: time.sleep(2 ** attempt)
            else: raise

def fetch_all(path, label):
    PAGE, rows, skip, pg = 500, [], 0, 1
    sep = "&" if "?" in path else "?"
    while True:
        print(f"  {label} p{pg} ({len(rows):,})...", flush=True)
        data = _get(f"{path}{sep}$top={PAGE}&$skip={skip}")
        page = data.get("value", data if isinstance(data, list) else [])
        rows.extend(page)
        if len(page) < PAGE: break
        skip += PAGE; pg += 1
    return rows

def df(field, frm, to):
    return f"{field} ge {frm}T00:00:00Z and {field} le {to}T23:59:59Z"

# ── 1. Fetch ──────────────────────────────────────────────────
print(f"Fetching Users...")
users = fetch_all("/Users", "users")
user_map = {u["ID"]: (u.get("FullName") or f"{u.get('FirstName','')} {u.get('LastName','')}".strip()) for u in users}

print(f"Fetching JobLines {YEAR_FROM} to {YEAR_TO}...")
f = urllib.parse.quote(df("OrderDate", YEAR_FROM, YEAR_TO))
lines = fetch_all(f"/JobLines?$filter={f}", "lines")
print(f"  {len(lines):,} lines fetched\n")

job_ids = list({l["JobID"] for l in lines if l.get("JobID")})
print(f"Fetching {len(job_ids):,} parent Jobs (chunked)...")
JOB_MAP, CHUNK = {}, 20
for i in range(0, len(job_ids), CHUNK):
    chunk = job_ids[i:i+CHUNK]
    fstr = urllib.parse.quote(" or ".join(f"ID eq {c}" for c in chunk))
    try:
        data = _get(f"/Jobs?$filter={fstr}&$top={CHUNK*2}")
        for j in data.get("value", []): JOB_MAP[j["ID"]] = j
    except Exception as e:
        print(f"  chunk {i} failed: {e}")
    if i % 1000 == 0: print(f"  jobs {min(i+CHUNK,len(job_ids)):,}/{len(job_ids):,}...", flush=True)
print(f"  {len(JOB_MAP):,} jobs loaded\n")

# ── 2. Classify lines (matches dashboard EXACTLY) ────────────
EXCLUDE_LINE_STATUS = {"status_job_line_cancelled"}
EXCLUDE_LINE_STAGE  = {"stage_job_line_unconfirmed"}
EXCLUDE_LINE_TYPE   = {"type_job_line_remake"}

def is_service_line(l, job):
    return (job or {}).get("JobType") == "type_job_service" or l.get("LineType") == "type_job_line_service"

def safe_disc(l):
    v = l.get("DiscountedPriceExTax")
    return v if v is not None else (l.get("StandardPriceExTax") or 0)

def disc_pct(std, disc):
    return max(0, min(100, (1 - disc/std) * 100)) if std > 0 else 0

def base_ref(ref):
    return re.sub(r"-\d+$", "", str(ref or ""))

non_cancelled = [l for l in lines if l.get("JobID") and l.get("Status") not in EXCLUDE_LINE_STATUS]
unconfirmed_lines = [l for l in non_cancelled if l.get("Stage") in EXCLUDE_LINE_STAGE]
remake_lines      = [l for l in non_cancelled if l.get("Stage") not in EXCLUDE_LINE_STAGE and l.get("LineType") in EXCLUDE_LINE_TYPE]
valid_lines       = [l for l in non_cancelled if l.get("Stage") not in EXCLUDE_LINE_STAGE and l.get("LineType") not in EXCLUDE_LINE_TYPE]

print(f"Core lines: {len(valid_lines):,}  |  Unconfirmed: {len(unconfirmed_lines):,}  |  Remake: {len(remake_lines):,}\n")

# ── 3. DisplayOption parsing ──────────────────────────────────
def parse_display_options(l):
    out = []
    for i in range(1, 11):
        raw = l.get(f"DisplayOption{i}")
        if not raw: continue
        idx = raw.find(":")
        if idx < 0: continue
        key = raw[:idx].strip()
        val = raw[idx+1:].strip()
        if not key: continue
        out.append((key, val or "(blank)"))
    return out

def fabric_of(l):
    for i in range(1, 11):
        raw = l.get(f"DisplayOption{i}")
        if raw and raw.lower().lstrip().startswith("fabric:"):
            return raw[raw.find(":")+1:].strip()
    return ""

# ── 4. Aggregates ─────────────────────────────────────────────
by_product = defaultdict(lambda: {"count":0, "qty":0, "revenue":0.0, "disc_sum":0.0, "suppliers":set(), "fabric_rev":defaultdict(float)})
by_supplier = defaultdict(lambda: {"count":0, "qty":0, "revenue":0.0})
by_attr = defaultdict(lambda: defaultdict(lambda: {"count":0, "qty":0, "revenue":0.0, "disc_sum":0.0}))
service_revenue, service_lines = 0.0, 0

for l in valid_lines:
    job = JOB_MAP.get(l["JobID"], {})
    std, disc = l.get("StandardPriceExTax") or 0, safe_disc(l)
    dpct = disc_pct(std, disc)
    qty = l.get("Qty") or 0

    p = by_product[l.get("Product") or "Unspecified"]
    p["count"] += 1; p["qty"] += qty; p["revenue"] += disc; p["disc_sum"] += dpct
    if l.get("Supplier"): p["suppliers"].add(l["Supplier"])
    fab = fabric_of(l)
    if fab: p["fabric_rev"][fab] += disc

    s = by_supplier[l.get("Supplier") or "Unspecified"]
    s["count"] += 1; s["qty"] += qty; s["revenue"] += disc

    for key, val in parse_display_options(l):
        a = by_attr[key][val]
        a["count"] += 1; a["qty"] += qty; a["revenue"] += disc; a["disc_sum"] += dpct

    if is_service_line(l, job):
        service_revenue += disc; service_lines += 1

fabric_rows = by_attr.get("Fabric", {})

unconfirmed_revenue = sum(safe_disc(l) for l in unconfirmed_lines)
remake_revenue = sum(safe_disc(l) for l in remake_lines)

# Attach rates
NEGATIVE_VALS = re.compile(r"^(none|no|n/a|nil|)$", re.I)
def attach_rate(key):
    data = by_attr.get(key, {})
    total = len(valid_lines)
    if not data or not total: return None
    positive = sum(v["count"] for val, v in data.items() if val and not NEGATIVE_VALS.match(val.strip()) and val != "(blank)")
    return positive / total * 100

ATTACH_KEYS = ["Motors", "Cassette", "Extension Brackets", "Bracket Covers"]

# ── 5. Build workbook ──────────────────────────────────────────
print("Building workbook...")
wb = openpyxl.Workbook()
wb.remove(wb.active)  # drop the default blank "Sheet"
NAV, WHT, PURPLE, SLATE_BG = "1E3A5F", "FFFFFF", "7C3AED", "F1F5F9"

def hdr_style(cell):
    cell.font = Font(name="Arial", bold=True, color=WHT, size=10)
    cell.fill = PatternFill("solid", fgColor=NAV)
    cell.alignment = Alignment(horizontal="left", vertical="center")

def body_style(cell, money=False, pct=False):
    cell.font = Font(name="Arial", size=10)
    cell.border = Border(bottom=Side(style="thin", color="E2E8F0"))
    cell.alignment = Alignment(horizontal="left", vertical="center")
    if money: cell.number_format = '$#,##0'
    if pct: cell.number_format = '0.0%'

def total_style(cell, money=False, pct=False):
    cell.font = Font(name="Arial", size=10, bold=True)
    cell.fill = PatternFill("solid", fgColor=SLATE_BG)
    cell.border = Border(top=Side(style="thin", color="94A3B8"))
    if money: cell.number_format = '$#,##0'
    if pct: cell.number_format = '0.0%'

def write_sheet(ws, headers, rows_data, col_widths, money_cols=(), pct_cols=(), sum_cols=(),
                 tab_color=None, total_row=True):
    """Returns (last_data_row, total_row_or_None) for precise cross-sheet formula refs."""
    if tab_color: ws.sheet_properties.tabColor = tab_color
    ws.append(headers)
    for cell in ws[1]: hdr_style(cell)
    for row in rows_data:
        ws.append(row)
        r = ws.max_row
        for ci in range(1, len(headers)+1):
            body_style(ws.cell(row=r, column=ci), money=ci in money_cols, pct=ci in pct_cols)
    ws.freeze_panes = "A2"
    last_data_row = ws.max_row
    if last_data_row > 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{last_data_row}"
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    tr = None
    if total_row and rows_data:
        tr = last_data_row + 1
        ws.cell(row=tr, column=1, value="TOTAL")
        for ci in sum_cols:
            col = get_column_letter(ci)
            ws.cell(row=tr, column=ci, value=f"=SUM({col}2:{col}{last_data_row})")
        for ci in range(1, len(headers)+1):
            total_style(ws.cell(row=tr, column=ci), money=ci in money_cols, pct=ci in pct_cols)
    return last_data_row, tr

# --- Sheet: Products ---
ws_p = wb.create_sheet("Products")
prod_rows = []
for name, p in sorted(by_product.items(), key=lambda x: -x[1]["revenue"]):
    top_fabric = max(p["fabric_rev"].items(), key=lambda x: x[1])[0] if p["fabric_rev"] else ""
    prod_rows.append([
        name, ", ".join(sorted(p["suppliers"])), p["count"], round(p["qty"],1),
        round(p["revenue"],2), round(p["revenue"]/p["count"],2),
        round(p["disc_sum"]/p["count"]/100, 4), top_fabric,
    ])
p_last, p_total = write_sheet(ws_p, ["Product","Supplier(s)","Lines","Units","Revenue","Avg Price","Avg Discount %","Top Fabric"],
            prod_rows, [32,30,8,9,14,12,13,32], money_cols=(5,6), pct_cols=(7,), sum_cols=(3,4,5), tab_color="0E9F8E")

# --- Sheet: Suppliers ---
ws_s = wb.create_sheet("Suppliers")
sup_rows = [[name, s["count"], round(s["qty"],1), round(s["revenue"],2)]
            for name, s in sorted(by_supplier.items(), key=lambda x: -x[1]["revenue"])]
s_last, s_total = write_sheet(ws_s, ["Supplier","Lines","Units","Revenue"], sup_rows, [32,10,10,14],
            money_cols=(4,), sum_cols=(2,3,4), tab_color="2E6DB4")

# --- Sheet: Fabrics ---
ws_f = wb.create_sheet("Fabrics")
fab_rows = [[fabric, v["count"], round(v["qty"],1), round(v["revenue"],2), round(v["disc_sum"]/v["count"]/100, 4)]
            for fabric, v in sorted(fabric_rows.items(), key=lambda x: -x[1]["revenue"])]
f_last, f_total = write_sheet(ws_f, ["Fabric","Lines","Units","Revenue","Avg Discount %"], fab_rows, [40,10,10,14,14],
            money_cols=(4,), pct_cols=(5,), sum_cols=(2,3,4), tab_color=PURPLE)

# --- Sheet: Attributes (all parts/options) — no single total row makes sense here,
# since each attribute key is its own dimension and summing across keys would multi-count lines.
ws_a = wb.create_sheet("Attributes")
attr_rows = []
for key in sorted(by_attr.keys(), key=lambda k: -sum(v["count"] for v in by_attr[k].values())):
    for val, v in sorted(by_attr[key].items(), key=lambda x: -x[1]["revenue"]):
        attr_rows.append([key, val, v["count"], round(v["qty"],1), round(v["revenue"],2), round(v["disc_sum"]/v["count"]/100, 4)])
write_sheet(ws_a, ["Attribute","Value","Lines","Units","Revenue","Avg Discount %"], attr_rows,
            [26,34,9,9,14,14], money_cols=(5,), pct_cols=(6,), tab_color="D97706", total_row=False)

# --- Sheet: Attach Rates ---
ws_ar = wb.create_sheet("Attach Rates")
ar_rows = []
for key in ATTACH_KEYS:
    rate = attach_rate(key)
    ar_rows.append([key, round(rate/100, 4) if rate is not None else None, sum(v["count"] for v in by_attr.get(key,{}).values())])
write_sheet(ws_ar, ["Part / Accessory","Attach Rate","Lines With This Option"], ar_rows, [26,14,20],
            pct_cols=(2,), tab_color=PURPLE, total_row=False)

# --- Sheet: Excluded (Unconfirmed & Remake) ---
ws_ex = wb.create_sheet("Excluded")
ex_rows = []
for l in unconfirmed_lines + remake_lines:
    job = JOB_MAP.get(l["JobID"], {})
    reason = "Unconfirmed Quote" if l.get("Stage") in EXCLUDE_LINE_STAGE else "Remake"
    ex_rows.append([
        job.get("Reference",""), base_ref(job.get("Reference") or str(l["JobID"])),
        (l.get("OrderDate") or "")[:10], user_map.get(job.get("SalesRepID"), ""),
        l.get("Product",""), reason, round(safe_disc(l),2),
    ])
ex_last, ex_total = write_sheet(ws_ex, ["Job Reference","Base Ref","Order Date","Consultant","Product","Reason Excluded","Revenue"],
            ex_rows, [16,14,12,22,32,18,12], money_cols=(7,), sum_cols=(7,), tab_color="64748B")

# --- Sheet: Raw Lines (source data) ---
ws_r = wb.create_sheet("Raw Lines")
raw_rows = []
for l in valid_lines:
    job = JOB_MAP.get(l["JobID"], {})
    std, disc = l.get("StandardPriceExTax") or 0, safe_disc(l)
    raw_rows.append([
        job.get("Reference",""), base_ref(job.get("Reference") or str(l["JobID"])),
        (l.get("OrderDate") or "")[:10], user_map.get(job.get("SalesRepID"), ""),
        l.get("Product",""), l.get("Supplier",""), fabric_of(l),
        l.get("Qty") or 0, round(std,2), round(disc,2), round(disc_pct(std,disc)/100, 4),
        "Yes" if is_service_line(l, job) else "No",
    ])
write_sheet(ws_r, ["Job Reference","Base Ref","Order Date","Consultant","Product","Supplier","Fabric",
                   "Qty","Standard Price","Sale Price (ex GST)","Discount %","Service?"],
            raw_rows, [16,14,12,22,32,20,32,7,14,16,12,10], money_cols=(9,10), pct_cols=(11,), total_row=False)

# --- Sheet: Summary (built last so it can reference the exact ranges above) ---
ws_sum = wb.create_sheet("Summary", 0)
wb.active = 0
ws_sum.sheet_properties.tabColor = NAV
ws_sum["A1"] = "Product Sales Analysis"
ws_sum["A1"].font = Font(name="Arial", bold=True, size=14, color=NAV)
ws_sum["A2"] = f"Period: {YEAR_FROM} to {YEAR_TO}  |  Generated: {TODAY.isoformat()}"
ws_sum["A2"].font = Font(name="Arial", italic=True, size=9, color="64748B")

summary_rows = [
    ("Lines Sold", f"=Products!C{p_total}", None, False),
    ("Units Sold", f"=Products!D{p_total}", None, False),
    ("Product Revenue (ex GST)", f"=Products!E{p_total}", "money", True),
    ("  of which Service", round(service_revenue, 2), "money", False),
    ("Avg Discount % (line-weighted)", f"=SUMPRODUCT(Products!C2:C{p_last},Products!G2:G{p_last})/Products!C{p_total}", "pct", False),
    ("Distinct Products", p_last - 1, None, False),
    ("Distinct Fabrics", f_last - 1, None, False),
    ("Distinct Suppliers", s_last - 1, None, False),
    ("", None, None, False),
    ("Excluded — Unconfirmed Quotes ($)", round(unconfirmed_revenue, 2), "money", True),
    ("Excluded — Unconfirmed Quotes (lines)", len(unconfirmed_lines), None, False),
    ("Excluded — Remake Lines ($)", round(remake_revenue, 2), "money", True),
    ("Excluded — Remake Lines (count)", len(remake_lines), None, False),
]

r0 = 4
ws_sum.cell(row=r0, column=1, value="Metric").font = Font(name="Arial", bold=True, size=10, color=WHT)
ws_sum.cell(row=r0, column=1).fill = PatternFill("solid", fgColor=NAV)
ws_sum.cell(row=r0, column=2, value="Value").font = Font(name="Arial", bold=True, size=10, color=WHT)
ws_sum.cell(row=r0, column=2).fill = PatternFill("solid", fgColor=NAV)
for i, (label, val, fmt, bold) in enumerate(summary_rows, start=r0+1):
    ws_sum.cell(row=i, column=1, value=label).font = Font(name="Arial", size=10, bold=bold)
    cell = ws_sum.cell(row=i, column=2, value=val)
    cell.font = Font(name="Arial", size=10, bold=bold)
    if fmt == "money": cell.number_format = '$#,##0'
    if fmt == "pct": cell.number_format = '0.0%'
ws_sum.column_dimensions["A"].width = 36
ws_sum.column_dimensions["B"].width = 20

note_row = r0 + len(summary_rows) + 2
ws_sum.cell(row=note_row, column=1, value=(
    "Note: Products/Suppliers/Fabrics/Attributes sheets are aggregated as of the export date above "
    "from the Raw Lines sheet. Excludes cancelled, unconfirmed-quote, and remake lines from the core "
    "total (tracked separately in the Excluded sheet); service revenue stays in the core total but is "
    "broken out above. Re-run export_product_sales_analysis.py for refreshed figures."
))
ws_sum.cell(row=note_row, column=1).font = Font(name="Arial", italic=True, size=8, color="94A3B8")
ws_sum.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=6)
ws_sum.cell(row=note_row, column=1).alignment = Alignment(wrap_text=True)

wb.save(OUT_FILE)
print(f"\n>> Saved {OUT_FILE}")
print(f"   Products: {len(prod_rows):,}  Suppliers: {len(sup_rows):,}  Fabrics: {len(fab_rows):,}  Attribute rows: {len(attr_rows):,}  Raw lines: {len(raw_rows):,}")
