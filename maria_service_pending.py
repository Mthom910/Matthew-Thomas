import urllib.request, urllib.parse, json, base64, time, re
from collections import Counter, defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from insyte_env import EMAIL, KEY
BASE  = "https://api.myinsyte.com.au/v2"
AUTH  = "Basic " + base64.b64encode(f"{EMAIL}:{KEY}".encode()).decode()
MARIA_ID = 54

def _get(path, retries=5):
    url = f"{BASE}/{path.lstrip('/')}"
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers={"Authorization": AUTH, "Accept": "application/json"})
            with urllib.request.urlopen(req, timeout=120) as r:
                return json.loads(r.read())
        except Exception as e:
            if attempt < retries - 1: time.sleep(2 ** attempt)
            else: raise

def fetch_all(path, label):
    PAGE, rows, skip, pg = 500, [], 0, 1
    sep = "&" if "?" in path else "?"
    while True:
        print(f"  {label} p{pg} ({len(rows)})...", flush=True)
        data = _get(f"{path}{sep}$top={PAGE}&$skip={skip}")
        page = data.get("value", data if isinstance(data, list) else [])
        rows.extend(page)
        if len(page) < PAGE: break
        skip += PAGE; pg += 1
    return rows

def base_ref(r):
    return re.sub(r'-\d+$', '', str(r or ''))

# ── 1. Fetch ALL of Maria's jobs (all stages, all types)
print("Fetching all of Maria's jobs...")
jf = urllib.parse.quote(f"SalesRepID eq {MARIA_ID}")
all_jobs = fetch_all(f"/Jobs?$filter={jf}", "Maria's jobs")
print(f"  {len(all_jobs):,} total jobs\n")

job_map = {j["ID"]: j for j in all_jobs}
job_ids = list(job_map.keys())

# ── 2. Fetch all job lines for Maria's jobs
print("Fetching job lines for all Maria's jobs...")
CHUNK = 15
all_lines = []
for i in range(0, len(job_ids), CHUNK):
    chunk = job_ids[i:i+CHUNK]
    fstr  = urllib.parse.quote(" or ".join(f"JobID eq {c}" for c in chunk))
    try:
        data = _get(f"/JobLines?$filter={fstr}&$top={CHUNK*20}")
        all_lines.extend(data.get("value", []))
    except Exception as e:
        print(f"  Warning chunk {i}: {e}")
    if i % 150 == 0:
        print(f"  lines {min(i+CHUNK, len(job_ids))}/{len(job_ids)}...", flush=True)

print(f"  {len(all_lines):,} total lines\n")

# ── 3. Check what DispatchStage values exist across Maria's lines
dispatch_vals = Counter(l.get("DispatchStage") for l in all_lines)
print("All DispatchStage values across Maria's lines:")
for k, v in dispatch_vals.most_common():
    print(f"  {v:4d}  {k!r}")

# ── 4. Filter to lines with service-pending dispatch stage
SERVICE_DISPATCH = {"dispatch_job_line_service_pending", "dispatch_job_line_servicepending"}
# Also catch any value containing "service"
service_lines = [l for l in all_lines
                 if (l.get("DispatchStage") or "").lower().replace("_","").replace(" ","")
                    in {"dispatchjoblineservicepending", "servicepending"}
                 or "service_pending" in (l.get("DispatchStage") or "").lower()
                 or "servicepending" in (l.get("DispatchStage") or "").lower().replace("_","")]

print(f"\nLines with service pending dispatch: {len(service_lines)}")

if not service_lines:
    # Broaden - any dispatch stage containing "service"
    service_lines = [l for l in all_lines if "service" in (l.get("DispatchStage") or "").lower()]
    print(f"Lines with 'service' in dispatch stage: {len(service_lines)}")
    for l in service_lines[:5]:
        print(f"  DispatchStage: {l.get('DispatchStage')!r}  Product: {l.get('Product')!r}")

# Group by base job reference
by_job = defaultdict(list)
for l in service_lines:
    jid = l.get("JobID")
    if jid:
        by_job[jid].append(l)

print(f"Unique jobs with service pending lines: {len(by_job)}\n")

# ── 5. Fetch contacts
contact_ids = list({job_map[jid].get("ContactID") for jid in by_job if job_map.get(jid, {}).get("ContactID")})
contacts = {}
for i in range(0, len(contact_ids), 20):
    chunk = contact_ids[i:i+20]
    fstr  = urllib.parse.quote(" or ".join(f"ID eq {c}" for c in chunk))
    try:
        data = _get(f"/Contacts?$filter={fstr}&$top=40")
        for c in data.get("value", []):
            contacts[c["ID"]] = c
    except: pass
print(f"  {len(contacts)} contacts fetched")

# ── 6. Fetch CM booking activities for these jobs
booking_dates = {}
jids_list = list(by_job.keys())
for i in range(0, len(jids_list), 15):
    chunk = jids_list[i:i+15]
    fstr  = urllib.parse.quote(" or ".join(f"JobID eq {c}" for c in chunk))
    try:
        data = _get(f"/Activities?$filter={fstr}&$top=150")
        for a in data.get("value", []):
            if a.get("Cancelled"): continue
            jid   = a.get("JobID")
            atype = (a.get("ActivityType") or "").lower()
            start = (a.get("Start") or "")[:10]
            if jid and start and ("cm" in atype or "book" in atype or "install" in atype or "service" in atype):
                if jid not in booking_dates or start > booking_dates[jid]["date"]:
                    booking_dates[jid] = {"date": start, "type": a.get("ActivityType")}
    except: pass

# ── 7. Build output rows
rows = []
for jid, lines in sorted(by_job.items()):
    job   = job_map.get(jid, {})
    ref   = base_ref(job.get("Reference", ""))
    full_ref = job.get("Reference", "")
    status = (job.get("Status") or "").replace("status_job_", "")
    stage  = (job.get("Stage") or "").replace("stage_job_", "")
    job_date = (job.get("JobDate") or "")[:10]
    eta_date = (job.get("ETADate") or "")[:10]

    cid   = job.get("ContactID")
    cont  = contacts.get(cid, {})
    cust  = f"{cont.get('FirstName','')} {cont.get('LastName','')}".strip() or f"Contact {cid}"
    mobile = cont.get("Mobile") or ""
    email  = cont.get("Email") or ""

    products = " | ".join({l.get("Product","") for l in lines if l.get("Product")})
    dispatch_stages = " | ".join({l.get("DispatchStage","") for l in lines if l.get("DispatchStage")})
    line_count = len(lines)

    bk = booking_dates.get(jid, {})
    booking_date = bk.get("date", "")
    booking_type = bk.get("type", "")

    rows.append({
        "Job Reference":    full_ref,
        "Base Ref":         ref,
        "Customer Name":    cust,
        "Mobile":           mobile,
        "Email":            email,
        "Job Date":         job_date,
        "Job Stage":        stage,
        "Job Status":       status,
        "Products":         products,
        "Dispatch Stage":   dispatch_stages,
        "Lines Pending":    line_count,
        "ETA Date":         eta_date,
        "Has Booking":      "Yes" if booking_date else "No",
        "Booking Date":     booking_date,
        "Booking Type":     booking_type,
        "Customer Notes":   job.get("CustomerNotes") or "",
        "Install Notes":    job.get("InstallNotes") or "",
    })

print(f"\nTotal jobs: {len(rows)}")
print(f"With booking: {sum(1 for r in rows if r['Has Booking']=='Yes')}")
print(f"Without booking: {sum(1 for r in rows if r['Has Booking']=='No')}")

# ── 8. Export to Excel
wb  = openpyxl.Workbook()
ws  = wb.active
ws.title = "Service Pending"

COLS = list(rows[0].keys()) if rows else []

hdr_fill  = PatternFill("solid", fgColor="1E3A5F")
hdr_font  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
body_font = Font(name="Arial", size=10)
green_fill = PatternFill("solid", fgColor="D1FAE5")
amber_fill = PatternFill("solid", fgColor="FEF3C7")
thin_border = Border(bottom=Side(style="thin", color="E2E8F0"))

ws.append(COLS)
for cell in ws[1]:
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal="left", vertical="center")

for row in rows:
    ws.append([row.get(c, "") for c in COLS])
    ri = ws.max_row
    bk_col = COLS.index("Has Booking") + 1
    cell = ws.cell(row=ri, column=bk_col)
    if cell.value == "Yes":
        cell.fill = green_fill
        cell.font = Font(name="Arial", size=10, bold=True, color="065F46")
    else:
        cell.fill = amber_fill
        cell.font = Font(name="Arial", size=10, bold=True, color="92400E")
    for c in ws[ri]:
        if c.column != bk_col:
            c.font = body_font
        c.border = thin_border
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

widths = {"Job Reference":16,"Base Ref":12,"Customer Name":22,"Mobile":15,"Email":28,
          "Job Date":12,"Job Stage":14,"Job Status":14,"Products":35,"Dispatch Stage":30,
          "Lines Pending":12,"ETA Date":12,"Has Booking":12,"Booking Date":14,
          "Booking Type":20,"Customer Notes":35,"Install Notes":35}
for i, col in enumerate(COLS, 1):
    ws.column_dimensions[get_column_letter(i)].width = widths.get(col, 16)

ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

out = "maria_rundle_service_pending.xlsx"
wb.save(out)
print(f"\n>> Saved: {out}")
